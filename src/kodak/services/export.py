"""Export range history (summary + raw transactions) to a single-sheet .xlsx."""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import select

from kodak import clock
from kodak.db import get_session
from kodak.models.cash import CashWithdrawal
from kodak.models.credit import Credit, CreditPayment
from kodak.models.enums import CreditStatus, ProductCategory
from kodak.models.price_history import ProductPriceHistory
from kodak.models.product import Product
from kodak.models.product_status_history import ProductStatusHistory
from kodak.models.user import User
from kodak.services.history import RangeSummary, TxnDetail, list_range_transactions

_CAT_LABEL: dict[ProductCategory, str] = {
    ProductCategory.photo:       "ფოტო",
    ProductCategory.enlargement: "გადიდება",
    ProductCategory.frame:       "ჩარჩო",
    ProductCategory.lamination:  "ლამინირება",
    ProductCategory.cd:          "CD",
    ProductCategory.photocopy:   "ქსეროქსი",
    ProductCategory.album:       "ალბომი",
    ProductCategory.other:       "სხვა",
}

_CREDIT_STATUS_LABEL: dict[CreditStatus, str] = {
    CreditStatus.active:   "გადაუხდელი",
    CreditStatus.partial:  "ნაწილობრივ გადახდილი",
    CreditStatus.cleared:  "დახურული",
    CreditStatus.forgiven: "ნაპატიები",
}

# ── styling tokens ────────────────────────────────────────────────────────
_ACCENT = "FF2E7D32"          # brand green
_ACCENT_SOFT = "FFE6F0E7"     # light green fill
_HEADER_TXT = "FFFFFFFF"
_BORDER_CLR = "FFD9D2C7"
_MONEY_FMT = '#,##0.00 ₾'

_TITLE_FONT   = Font(size=16, bold=True, color=_ACCENT)
_SECTION_FONT = Font(size=12, bold=True, color=_ACCENT)
_LABEL_FONT   = Font(size=11)
_VALUE_FONT   = Font(size=11, bold=True)
_HEAD_FONT    = Font(size=11, bold=True, color=_HEADER_TXT)

_HEAD_FILL    = PatternFill("solid", fgColor=_ACCENT)
_SOFT_FILL    = PatternFill("solid", fgColor=_ACCENT_SOFT)

_THIN = Side(style="thin", color=_BORDER_CLR)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_RIGHT = Alignment(horizontal="right", vertical="center")
_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# Transaction column layout: (header, width, alignment)
_TXN_COLUMNS: list[tuple[str, int, Alignment]] = [
    ("თარიღი",     12, _LEFT),
    ("დრო",         8, _LEFT),
    ("გვარი",      18, _LEFT),
    ("პროდუქტები", 46, _WRAP),
    ("ჯამი",       12, _RIGHT),
    ("გადახდილი",  12, _RIGHT),
    ("ნისია",      12, _RIGHT),
    ("სტატუსი",    12, _LEFT),
    ("შენიშვნა",   28, _WRAP),
]


def export_history_to_xlsx(
    path: Path,
    *,
    start: dt.date,
    end: dt.date,
    summary: RangeSummary,
    rows: list[TxnDetail],
) -> Path:
    """Write a single-sheet workbook: summary metrics on top, raw txns below.

    Returns the written path. Raises on I/O failure (caller handles feedback).
    """
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "ისტორია"
    ws.sheet_view.showGridLines = False

    _apply_column_widths(ws)

    row = _write_summary_block(ws, start, end, summary)
    row += 1  # spacer row
    _write_transactions_block(ws, rows, start_row=row)

    ws.freeze_panes = None
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


# ── summary ────────────────────────────────────────────────────────────────

def _write_summary_block(
    ws: Worksheet, start: dt.date, end: dt.date, s: RangeSummary
) -> int:
    """Write the header + metric block. Returns the next free row index."""
    r = 1
    cell = ws.cell(row=r, column=1, value="გაყიდვების ისტორია — ანგარიში")
    cell.font = _TITLE_FONT
    r += 1

    period = (
        f"{start:%d.%m.%Y}"
        if start == end
        else f"{start:%d.%m.%Y} – {end:%d.%m.%Y}"
    )
    _kv(ws, r, "პერიოდი:", period)
    r += 1
    _kv(ws, r, "შექმნის თარიღი:", f"{clock.now():%d.%m.%Y %H:%M}")
    r += 2

    ws.cell(row=r, column=1, value="შემაჯამებელი მაჩვენებლები").font = _SECTION_FONT
    r += 1

    metrics: list[tuple[str, object, bool]] = [
        ("სულ გაყიდვები",         s.total_txns,              False),
        ("ჯამური გაყიდვები",      s.total_revenue,           True),
        ("მიღებული გაყიდვებიდან", s.received_from_sales,     True),
        ("დაბრუნებული ნისიები",   s.credit_received_amount,  True),
        ("ახალი ნისიები",         s.new_credit_count,        False),
        ("სალაროში მიღებული",     s.cashier_received,        True),
    ]
    for label, value, is_money in metrics:
        _metric_row(ws, r, label, value, money=is_money)
        r += 1

    # Category breakdown (part of the on-screen summary).
    if s.categories:
        r += 1
        ws.cell(row=r, column=1, value="კატეგორიები").font = _SECTION_FONT
        r += 1
        for col, head in enumerate(("კატეგორია", "რაოდენობა", "შემოსავალი"), start=1):
            c = ws.cell(row=r, column=col, value=head)
            c.font = _HEAD_FONT
            c.fill = _HEAD_FILL
            c.border = _CELL_BORDER
            c.alignment = _LEFT if col == 1 else _RIGHT
        r += 1
        for cs in s.categories:
            name = _CAT_LABEL.get(cs.category, cs.category.value)
            ws.cell(row=r, column=1, value=name).border = _CELL_BORDER
            qc = ws.cell(row=r, column=2, value=cs.qty)
            qc.border = _CELL_BORDER
            qc.alignment = _RIGHT
            rc = ws.cell(row=r, column=3, value=_money(cs.revenue))
            rc.number_format = _MONEY_FMT
            rc.border = _CELL_BORDER
            rc.alignment = _RIGHT
            r += 1

    return r


def _kv(ws: Worksheet, r: int, label: str, value: str) -> None:
    a = ws.cell(row=r, column=1, value=label)
    a.font = _LABEL_FONT
    b = ws.cell(row=r, column=2, value=value)
    b.font = _VALUE_FONT


def _metric_row(ws: Worksheet, r: int, label: str, value, *, money: bool) -> None:
    a = ws.cell(row=r, column=1, value=label)
    a.font = _LABEL_FONT
    a.fill = _SOFT_FILL
    a.border = _CELL_BORDER
    a.alignment = _LEFT
    b = ws.cell(row=r, column=2, value=_money(value) if money else value)
    b.font = _VALUE_FONT
    b.fill = _SOFT_FILL
    b.border = _CELL_BORDER
    b.alignment = _RIGHT
    if money:
        b.number_format = _MONEY_FMT


# ── transactions ─────────────────────────────────────────────────────────

def _write_transactions_block(
    ws: Worksheet, rows: list[TxnDetail], *, start_row: int
) -> int:
    r = start_row
    ws.cell(row=r, column=1, value="ტრანზაქციები").font = _SECTION_FONT
    r += 1

    header_row = r
    for col, (head, _width, _align) in enumerate(_TXN_COLUMNS, start=1):
        c = ws.cell(row=header_row, column=col, value=head)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _CELL_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    if not rows:
        ws.cell(row=r, column=1, value="ამ პერიოდში ჩანაწერი არ არის.").font = _LABEL_FONT
        return r + 1

    # Newest first, matching the on-screen list.
    ordered = sorted(rows, key=lambda d: (d.txn.date, d.txn.created_at), reverse=True)
    for d in ordered:
        _write_txn_row(ws, r, d)
        r += 1

    return r


def _write_txn_row(ws: Worksheet, r: int, d: TxnDetail) -> None:
    products = "; ".join(
        f"{prod.name} {prod.size_label or ''}".strip() + f" ×{li.quantity}"
        for li, prod in d.items
    )

    if d.credit is not None:
        credit_remaining = _money(d.credit.remaining_amount)
        credit_status = _CREDIT_STATUS_LABEL.get(d.credit.status, d.credit.status.value)
    else:
        credit_remaining = None
        credit_status = "—"

    values = [
        f"{d.txn.date:%d.%m.%Y}",
        clock.to_local(d.txn.created_at).strftime("%H:%M"),
        d.txn.customer_surname,
        products or "—",
        _money(d.total),
        _money(d.txn.amount_received),
        credit_remaining,
        credit_status,
        d.txn.notes or "",
    ]
    for col, ((_head, _width, align), value) in enumerate(zip(_TXN_COLUMNS, values), start=1):
        c = ws.cell(row=r, column=col, value=value)
        c.border = _CELL_BORDER
        c.alignment = align
        c.font = _LABEL_FONT
        # Money columns: ჯამი, გადახდილი, ნისია → indexes 5,6,7
        if col in (5, 6, 7) and value is not None:
            c.number_format = _MONEY_FMT


# ── helpers ──────────────────────────────────────────────────────────────

def _apply_column_widths(ws: Worksheet) -> None:
    for col, (_head, width, _align) in enumerate(_TXN_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _money(value) -> float:
    if isinstance(value, Decimal):
        return float(value)
    if value is None:
        return 0.0
    return float(value)


# ── full export (all data, multi-sheet) ────────────────────────────────────

def export_all_to_xlsx(path: Path) -> Path:
    """Export the whole database to a multi-sheet workbook.

    Sheets: გაყიდვები (transactions), ნისია (credits), ნისიის გადახდები
    (credit repayments), გატანები (cash withdrawals), პროდუქტები (price list).
    """
    path = Path(path)
    with get_session() as session:
        txns = list_range_transactions(session, dt.date(2000, 1, 1), clock.today())
        credits = list(session.exec(select(Credit).order_by(Credit.date)).all())
        payments = list(
            session.exec(select(CreditPayment).order_by(CreditPayment.date)).all()
        )
        withdrawals = list(
            session.exec(select(CashWithdrawal).order_by(CashWithdrawal.date)).all()
        )
        products = list(
            session.exec(
                select(Product).order_by(Product.category, Product.sort_order)
            ).all()
        )
        users = {u.id: u for u in session.exec(select(User)).all()}
        credit_code = {c.id: c.code for c in credits}

    def _user_name(uid) -> str:
        u = users.get(uid)
        return u.full_name if u else "—"

    wb = Workbook()

    # 1) Transactions
    ws = wb.active
    ws.title = "გაყიდვები"
    _fill_transactions_sheet(ws, txns)

    # 2) Credits
    _write_table(
        wb.create_sheet("ნისია"),
        [
            ("კოდი", 14, _LEFT), ("თარიღი", 12, _LEFT), ("გვარი", 18, _LEFT),
            ("საწყისი", 12, _RIGHT), ("დარჩენილი", 12, _RIGHT),
            ("სტატუსი", 12, _LEFT), ("ნაპატიების დრო", 18, _LEFT),
        ],
        [
            [
                c.code, f"{c.date:%d.%m.%Y}", c.customer_surname,
                _money(c.original_amount), _money(c.remaining_amount),
                _CREDIT_STATUS_LABEL.get(c.status, c.status.value),
                (clock.to_local(c.forgiven_at).strftime("%d.%m.%Y %H:%M")
                 if c.forgiven_at else ""),
            ]
            for c in credits
        ],
        money_cols={4, 5},
    )

    # 3) Credit repayments
    _write_table(
        wb.create_sheet("ნისიის გადახდები"),
        [
            ("თარიღი", 12, _LEFT), ("ნისიის კოდი", 14, _LEFT),
            ("თანხა", 12, _RIGHT), ("თანამშრომელი", 18, _LEFT),
            ("შენიშვნა", 28, _WRAP),
        ],
        [
            [
                f"{p.date:%d.%m.%Y}", credit_code.get(p.credit_id, "—"),
                _money(p.amount), _user_name(p.created_by_user_id), p.notes or "",
            ]
            for p in payments
        ],
        money_cols={3},
    )

    # 4) Cash withdrawals
    _write_table(
        wb.create_sheet("გატანები"),
        [
            ("თარიღი", 12, _LEFT), ("დრო", 8, _LEFT),
            ("თანამშრომელი", 18, _LEFT), ("თანხა", 12, _RIGHT),
            ("შენიშვნა", 28, _WRAP),
        ],
        [
            [
                f"{w.date:%d.%m.%Y}",
                clock.to_local(w.created_at).strftime("%H:%M"),
                _user_name(w.user_id), _money(w.amount), w.note or "",
            ]
            for w in withdrawals
        ],
        money_cols={4},
    )

    # 5) Products (price list)
    _write_table(
        wb.create_sheet("პროდუქტები"),
        [
            ("კატეგორია", 16, _LEFT), ("დასახელება", 26, _LEFT),
            ("ზომა", 12, _LEFT), ("ფასი", 12, _RIGHT), ("აქტიური", 10, _LEFT),
        ],
        [
            [
                _CAT_LABEL.get(p.category, p.category.value), p.name,
                p.size_label or "", _money(p.unit_price),
                "კი" if p.active else "არა",
            ]
            for p in products
        ],
        money_cols={4},
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def export_products_group_to_xlsx(
    path: Path, *, category: ProductCategory | None = None
) -> Path:
    """Export current prices and price/status changes for one product group."""
    path = Path(path)
    with get_session() as session:
        query = select(Product).order_by(Product.sort_order, Product.name)
        if category is not None:
            query = query.where(Product.category == category)
        products = list(session.exec(query).all())
        product_ids = [p.id for p in products if p.id is not None]

        price_logs: list[ProductPriceHistory] = []
        status_logs: list[ProductStatusHistory] = []
        if product_ids:
            price_logs = list(
                session.exec(
                    select(ProductPriceHistory)
                    .where(ProductPriceHistory.product_id.in_(product_ids))
                    .order_by(ProductPriceHistory.changed_at)
                ).all()
            )
            status_logs = list(
                session.exec(
                    select(ProductStatusHistory)
                    .where(ProductStatusHistory.product_id.in_(product_ids))
                    .order_by(ProductStatusHistory.changed_at)
                ).all()
            )

        user_ids = {
            row.changed_by_user_id
            for row in [*price_logs, *status_logs]
            if row.changed_by_user_id
        }
        users: dict[int, User] = {}
        if user_ids:
            users = {
                u.id: u
                for u in session.exec(select(User).where(User.id.in_(user_ids))).all()
            }

    product_by_id = {p.id: p for p in products if p.id is not None}
    group_name = "ყველა ჯგუფი" if category is None else _CAT_LABEL.get(category, category.value)

    def product_label(product_id: int) -> str:
        product = product_by_id.get(product_id)
        if product is None:
            return "—"
        return f"{product.name} {product.size_label or ''}".strip()

    def user_name(user_id: int | None) -> str:
        user = users.get(user_id)
        return user.full_name if user else "—"

    wb = Workbook()
    ws = wb.active
    ws.title = "მიმდინარე ფასები"
    _write_table(
        ws,
        [
            ("ჯგუფი", 16, _LEFT), ("დასახელება", 28, _LEFT),
            ("ზომა", 14, _LEFT), ("ფასი", 12, _RIGHT),
            ("სტატუსი", 12, _LEFT), ("განახლდა", 18, _LEFT),
        ],
        [
            [
                _CAT_LABEL.get(p.category, p.category.value),
                p.name,
                p.size_label or "",
                _money(p.unit_price),
                _active_label(p.active),
                clock.to_local(p.updated_at).strftime("%d.%m.%Y %H:%M"),
            ]
            for p in products
        ],
        money_cols={4},
    )
    ws.cell(row=1, column=8, value="ჯგუფი:").font = _LABEL_FONT
    ws.cell(row=1, column=9, value=group_name).font = _VALUE_FONT

    log_rows: list[tuple[dt.datetime, list]] = []
    for row in price_logs:
        log_rows.append((
            row.changed_at,
            [
                clock.to_local(row.changed_at).strftime("%d.%m.%Y"),
                clock.to_local(row.changed_at).strftime("%H:%M"),
                product_label(row.product_id),
                "ფასი",
                f"₾{_money(row.old_price):.2f}",
                f"₾{_money(row.new_price):.2f}",
                user_name(row.changed_by_user_id),
            ],
        ))
    for row in status_logs:
        log_rows.append((
            row.changed_at,
            [
                clock.to_local(row.changed_at).strftime("%d.%m.%Y"),
                clock.to_local(row.changed_at).strftime("%H:%M"),
                product_label(row.product_id),
                "სტატუსი",
                _active_label(row.old_active),
                _active_label(row.new_active),
                user_name(row.changed_by_user_id),
            ],
        ))

    _write_table(
        wb.create_sheet("ცვლილებების ჟურნალი"),
        [
            ("თარიღი", 12, _LEFT), ("დრო", 8, _LEFT),
            ("პროდუქტი", 30, _LEFT), ("ცვლილება", 12, _LEFT),
            ("ძველი", 14, _LEFT), ("ახალი", 14, _LEFT),
            ("მომხმარებელი", 18, _LEFT),
        ],
        [row for _changed_at, row in sorted(log_rows, key=lambda item: item[0])],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _fill_transactions_sheet(ws: Worksheet, rows: list[TxnDetail]) -> None:
    ws.sheet_view.showGridLines = False
    _apply_column_widths(ws)
    for col, (head, _width, _align) in enumerate(_TXN_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=head)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _CELL_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    r = 2
    for d in sorted(rows, key=lambda d: (d.txn.date, d.txn.created_at)):
        _write_txn_row(ws, r, d)
        r += 1


def _write_table(
    ws: Worksheet,
    headers: list[tuple[str, int, Alignment]],
    rows: list[list],
    *,
    money_cols: set[int] = frozenset(),
) -> None:
    """Generic single-table sheet: bold header band + bordered data rows."""
    ws.sheet_view.showGridLines = False
    for col, (label, width, _align) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _CELL_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    for ri, row in enumerate(rows, start=2):
        for col, (value, (_label, _width, align)) in enumerate(
            zip(row, headers), start=1
        ):
            c = ws.cell(row=ri, column=col, value=value)
            c.border = _CELL_BORDER
            c.alignment = align
            c.font = _LABEL_FONT
            if col in money_cols and value is not None:
                c.number_format = _MONEY_FMT


def _active_label(active: bool) -> str:
    return "აქტიური" if active else "გამორთული"
